# -*- coding: utf-8 -*-
"""
生成 NMR 核磁测试登记表 — 批量导入模板
输出：NMR批量导入模板.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ 样式定义 ============
HEADER_FILL = PatternFill(start_color="4a148c", end_color="4a148c", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
EXAMPLE_FONT = Font(name="微软雅黑", size=10, color="555555")
EXAMPLE_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="4a148c")
SUB_FONT = Font(name="微软雅黑", size=10, color="666666")
NOTE_FONT = Font(name="微软雅黑", size=10, color="333333")
SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="4a148c")
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============ Sheet1: 导入数据 ============
ws = wb.active
ws.title = "导入数据"

# 列定义（顺序与系统导入逻辑一致）
COLUMNS = [
    ("登记日期", 14, "格式：2026-06-20 或 2026/6/20"),
    ("样品编号", 14, "如：WK-001"),
    ("送样人", 10, "如：李俊凯"),
    ("单位/部门", 16, "如：长江大学"),
    ("样品状态", 12, "固体 / 液体 / 溶解液体"),
    ("保存条件", 10, "室温 / 冷藏 / 冷冻"),
    ("溶剂/氘代试剂", 24, "见说明sheet，如：氘代氯仿(CDCl₃)"),
    ("测试项目", 28, "多个用逗号分隔，如：1H谱(16ns),13C谱(512ns)"),
    ("优先级", 8, "正常 / 加急"),
    ("状态", 10, "待测试 / 测试中 / 已测试"),
    ("付款状态", 10, "未付款 / 已付款"),
    ("备注", 20, "可选，支持费用调整如：加急[+10元]"),
]

# 写表头
for idx, (name, width, _hint) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=idx, value=name)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(idx)].width = width
ws.row_dimensions[1].height = 28

# 写示例数据（3行）
EXAMPLES = [
    ["2026-06-20", "WK-001", "李俊凯", "长江大学", "固体", "室温", "氘代氯仿(CDCl₃)", "1H谱(16ns),13C谱(512ns)", "正常", "已测试", "未付款", "常规样品"],
    ["2026-06-20", "WK-002", "徐志红", "辽宁科技学院", "液体", "冷藏", "氘代二甲亚砜(DMSO-d₆)", "1H谱(16ns)", "加急", "测试中", "未付款", "加急处理[+10元]"],
    ["2026-06-21", "WK-003", "姜春风", "长江大学", "溶解液体", "室温", "重水(D₂O)", "COSY (64ns),HSQC (512ns)", "正常", "待测试", "已付款", "二维谱测试"],
]
for r, row_data in enumerate(EXAMPLES, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = EXAMPLE_FONT
        cell.fill = EXAMPLE_FILL
        cell.alignment = left_align
        cell.border = thin_border
    ws.row_dimensions[r].height = 22

# 预留空行（带边框，方便填写）
for r in range(5, 25):
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = left_align
        cell.font = Font(name="微软雅黑", size=10)
    ws.row_dimensions[r].height = 20

# ============ 数据验证（下拉框）============
# E列：样品状态
dv_status = DataValidation(type="list", formula1='"固体,液体,溶解液体"', allow_blank=True)
dv_status.error = "请选择：固体 / 液体 / 溶解液体"
dv_status.errorTitle = "样品状态无效"
ws.add_data_validation(dv_status)
dv_status.add("E2:E1000")

# F列：保存条件
dv_storage = DataValidation(type="list", formula1='"室温,冷藏,冷冻"', allow_blank=True)
ws.add_data_validation(dv_storage)
dv_storage.add("F2:F1000")

# G列：溶剂
SOLVENTS = [
    "无溶剂", "氘代氯仿(CDCl₃)", "氘代二甲亚砜(DMSO-d₆)", "重水(D₂O)",
    "氘代丙酮(C₃D₆O)", "氘代乙腈(CD₃CN-d₃)", "氘代乙酸(CD₃COOH-d₄)",
    "氘代盐酸(DCl)", "氘代吡啶(Py-d₅)", "氘代三氟乙酸(CF₃COOD)"
]
# Excel公式字符串长度限制，用命名区域或直接列表
solvent_formula = '"' + ",".join(SOLVENTS) + '"'
dv_solvent = DataValidation(type="list", formula1=solvent_formula, allow_blank=True)
ws.add_data_validation(dv_solvent)
dv_solvent.add("G2:G1000")

# I列：优先级
dv_priority = DataValidation(type="list", formula1='"正常,加急"', allow_blank=True)
ws.add_data_validation(dv_priority)
dv_priority.add("I2:I1000")

# J列：状态
dv_test_status = DataValidation(type="list", formula1='"待测试,测试中,已测试"', allow_blank=True)
ws.add_data_validation(dv_test_status)
dv_test_status.add("J2:J1000")

# K列：付款状态
dv_payment = DataValidation(type="list", formula1='"未付款,已付款"', allow_blank=True)
ws.add_data_validation(dv_payment)
dv_payment.add("K2:K1000")

# 冻结首行
ws.freeze_panes = "A2"

# ============ Sheet2: 填写说明 ============
ws2 = wb.create_sheet("填写说明")
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 12

# 标题
ws2.merge_cells("A1:D1")
t = ws2["A1"]
t.value = "NMR 核磁测试登记表 — 批量导入模板填写说明"
t.font = TITLE_FONT
t.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:D2")
s = ws2["A2"]
s.value = "请在「导入数据」sheet 中填写数据，保存后通过系统「批量导入」按钮上传 CSV/Excel 文件"
s.font = SUB_FONT
s.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[2].height = 22

# ---- 字段说明表 ----
row = 4
ws2.cell(row=row, column=1, value="一、字段说明").font = SECTION_FONT
row += 1
headers = ["列名", "是否必填", "可选值/格式", "说明"]
for c, h in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = center_align
    cell.border = thin_border
ws2.row_dimensions[row].height = 26

FIELD_DOCS = [
    ("登记日期", "必填", "YYYY-MM-DD（如2026-06-20）", "支持 2026/6/20、06/20/2026 等格式，系统自动转换"),
    ("样品编号", "必填", "自由文本（如WK-001）", "每条记录唯一标识，建议有规律便于查询"),
    ("送样人", "必填", "自由文本", "如：李俊凯、徐志红、姜春风"),
    ("单位/部门", "必填", "自由文本", "如：长江大学、辽宁科技学院"),
    ("样品状态", "必填", "固体 / 液体 / 溶解液体", "下拉选择"),
    ("保存条件", "必填", "室温 / 冷藏 / 冷冻", "下拉选择"),
    ("溶剂/氘代试剂", "必填", "见下方溶剂价格表", "下拉选择，支持简写如 CDCl₃、DMSO"),
    ("测试项目", "必填", "见下方测试项目价格表", "多个项目用英文逗号分隔，如：1H谱(16ns),13C谱(512ns)"),
    ("优先级", "选填", "正常 / 加急（默认正常）", "下拉选择"),
    ("状态", "选填", "待测试 / 测试中 / 已测试（默认已测试）", "下拉选择"),
    ("付款状态", "选填", "未付款 / 已付款（默认未付款）", "下拉选择"),
    ("备注", "选填", "自由文本", "支持费用调整标记，格式：备注内容[+10元]或[-5元折扣]"),
]
for doc in FIELD_DOCS:
    row += 1
    for c, val in enumerate(doc, 1):
        cell = ws2.cell(row=row, column=c, value=val)
        cell.font = NOTE_FONT
        cell.alignment = left_align
        cell.border = thin_border
    ws2.row_dimensions[row].height = 22

# ---- 溶剂价格表 ----
row += 2
ws2.cell(row=row, column=1, value="二、溶剂/氘代试剂价格表").font = SECTION_FONT
row += 1
for c, h in enumerate(["溶剂名称", "费用", "支持简写", ""], 1):
    cell = ws2.cell(row=row, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = center_align
    cell.border = thin_border
ws2.row_dimensions[row].height = 26

SOLVENT_TABLE = [
    ("无溶剂", "0元（免费）", ""),
    ("氘代氯仿(CDCl₃)", "0元（免费）", "CDCl₃ / 氘代氯仿 / Chloroform"),
    ("氘代二甲亚砜(DMSO-d₆)", "0元（免费）", "DMSO / DMSO-d₆"),
    ("重水(D₂O)", "0元（免费）", "D₂O / 重水"),
    ("氘代丙酮(C₃D₆O)", "25元", "氘代丙酮 / C₃D₆O / Acetone"),
    ("氘代乙腈(CD₃CN-d₃)", "25元", "氘代乙腈 / CD₃CN / Acetonitrile"),
    ("氘代乙酸(CD₃COOH-d₄)", "30元", "氘代乙酸 / CD₃COOH"),
    ("氘代盐酸(DCl)", "30元", "DCl / 氘代盐酸"),
    ("氘代吡啶(Py-d₅)", "70元", "氘代吡啶 / Py-d₅ / Pyridine"),
    ("氘代三氟乙酸(CF₃COOD)", "35元", "TFA / CF₃COOD"),
]
for s in SOLVENT_TABLE:
    row += 1
    for c, val in enumerate(s, 1):
        cell = ws2.cell(row=row, column=c, value=val)
        cell.font = NOTE_FONT
        cell.alignment = left_align
        cell.border = thin_border
    ws2.row_dimensions[row].height = 20

# ---- 测试项目价格表 ----
row += 2
ws2.cell(row=row, column=1, value="三、测试项目价格表").font = SECTION_FONT
row += 1
for c, h in enumerate(["测试项目", "费用（元/样）", "支持简写", ""], 1):
    cell = ws2.cell(row=row, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = center_align
    cell.border = thin_border
ws2.row_dimensions[row].height = 26

TEST_TABLE = [
    ("1H谱(16ns)", "20", "1H / 1H谱 / HNMR / 1H-NMR"),
    ("13C谱(512ns)", "40", "13C / 13C谱 / CNMR / 13C-NMR"),
    ("13C谱(1024ns)", "60", "（注意与512ns区分）"),
    ("19F谱(64ns)", "25", "19F / 19F谱 / 19F-NMR"),
    ("31P谱(64ns)", "30", "31P / 31P谱 / 31P-NMR"),
    ("11B谱(64ns)", "30", "11B / 11B谱 / 11B-NMR"),
    ("DEPT45(512ns)", "40", "DEPT45"),
    ("DEPT90(512ns)", "40", "DEPT90"),
    ("DEPT135(512ns)", "40", "DEPT135 / DEPT-135"),
    ("COSY (64ns)", "35", "COSY"),
    ("HSQC (512ns)", "40", "HSQC"),
    ("NOESY (H-H)", "90", "NOESY"),
    ("HMBC (H-C)", "100", "HMBC"),
]
for t in TEST_TABLE:
    row += 1
    for c, val in enumerate(t, 1):
        cell = ws2.cell(row=row, column=c, value=val)
        cell.font = NOTE_FONT
        cell.alignment = left_align
        cell.border = thin_border
    ws2.row_dimensions[row].height = 20

# ---- 费用调整说明 ----
row += 2
ws2.cell(row=row, column=1, value="四、备注列费用调整格式").font = SECTION_FONT
row += 1
ADJUST_DOCS = [
    ("加费", "加急处理[+10元]", "在备注末尾加 [+金额元]，总费用增加"),
    ("折扣", "内部样品[-5元折扣]", "在备注末尾加 [-金额元折扣]，总费用减少"),
    ("无调整", "常规样品", "不加标记，费用按标准计算"),
]
for c, h in enumerate(["类型", "示例", "说明"], 1):
    cell = ws2.cell(row=row, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = center_align
    cell.border = thin_border
ws2.row_dimensions[row].height = 26
for a in ADJUST_DOCS:
    row += 1
    for c, val in enumerate(a, 1):
        cell = ws2.cell(row=row, column=c, value=val)
        cell.font = NOTE_FONT
        cell.alignment = left_align
        cell.border = thin_border
    ws2.row_dimensions[row].height = 22

# ---- 导入步骤 ----
row += 2
ws2.cell(row=row, column=1, value="五、导入步骤").font = SECTION_FONT
STEPS = [
    "1. 在「导入数据」sheet 中填写数据（可删除示例行，保留表头）",
    "2. 将文件另存为 CSV（逗号分隔）格式，编码选 UTF-8",
    "3. 打开 NMR 核磁测试登记表系统 → 点击「批量导入」按钮",
    "4. 选择 CSV 文件 或 直接粘贴 Excel 数据（Tab分隔）到文本框",
    "5. 点击「导入」按钮，系统自动解析并添加记录",
    "",
    "提示：直接从 Excel 复制单元格（含表头），粘贴到导入文本框也可识别",
    "提示：日期格式支持多种写法，系统自动转换；溶剂和测试项目支持简写",
]
for step in STEPS:
    row += 1
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws2.cell(row=row, column=1, value=step)
    cell.font = NOTE_FONT
    cell.alignment = left_align
    ws2.row_dimensions[row].height = 20

# 保存
output = "public/NMR批量导入模板.xlsx"
wb.save(output)
print(f"✅ Excel模板已生成: {output}")
print(f"   Sheet1: 导入数据（含表头+示例+下拉验证）")
print(f"   Sheet2: 填写说明（字段说明+价格表+导入步骤）")
